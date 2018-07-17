from openpyxl import load_workbook
import openpyxl.styles as sty
import datetime
import csv


def get_map(path):
    wechat2name = {}
    qq2name = {}
    workbook = load_workbook(path)
    sheets = workbook.sheetnames
    booksheet = workbook[sheets[0]]
    rows = booksheet.rows
    for row in rows:
        line = [col.value for col in row]
        if line[1] and line[1].startswith("o86"):
            wechat2name[line[1]] = line[2]
        if line[0] and isinstance(line[0],int):
            qq2name[line[0]] = line[2]
    return wechat2name, qq2name


def check_talbe(path,wechat2name,qq2name):
    name_sets = {'A':set(),'B':set(),'C':set()}
    with open(path,'r') as fin, open("clock_in.log","w") as log:
        rows = csv.reader(fin, delimiter=',')
        next(rows)
        for line in rows:
            if len(line[7]) > 1:
                if len(line[7]) < 5:
                    name_sets[line[6][0]].add(line[7])
                else:
                    print("name:", line[7], file=log)
            if line[4].startswith("o86"):
                if line[4] in wechat2name:
                    name = wechat2name[line[4]]
                    name_sets[line[6][0]].add(name)
                else:
                    print("wechat:", line[4], file=log)
            if len(line[3]) > 1:
                qq = int(line[3])
                if qq in qq2name:
                    name = qq2name[qq]
                    name_sets[line[6][0]].add(name)
                else:
                    print("qq:", qq, file=log)
    return name_sets


def clock_in(path, name_sets):
    workbook = load_workbook(path)
    sheets = workbook.sheetnames
    booksheet = workbook[sheets[1]]
    col = next(booksheet.columns)
    name_list = [cell.value for cell in col]
    name2room = {}
    room2names = {}
    cnt = 0
    for col in booksheet.columns:
        cnt += 1
        if cnt == 3:
            for i, each in enumerate(col):
                name = name_list[i]
                room = each.value
                name2room[name] = room
                if room not in room2names:
                    room2names[room] = set()
                room2names[room].add(name)
    for i, each in enumerate(col):
        if each.value == "已回家":
            name_sets['C'].add(name_list[i])
    print("不顺利",name_sets['B'])
    print("已回家",name_sets['C'])
    not_clock_in = set(name_list[1:]).difference(name_sets['A'].union(name_sets['B']).union(name_sets['C']))
    print("未签到")
    undo = []
    for each in not_clock_in:
        flag = True
        print(" + "+each,end=" : ")
        room = name2room[each]
        names = room2names[room]
        for name in names:
            if name != each and name not in not_clock_in:
                print(name,end=" ")
                flag = False
        if flag:
            undo.append(each)
        print()
    day_cell = booksheet.cell(row=1, column=cnt)
    cnt += 1
    new_day_cell = booksheet.cell(row=1, column=cnt)
    new_day_cell.value = day_cell.value + datetime.timedelta(days = 1)
    new_day_cell.number_format = day_cell.number_format
    for i, name in enumerate(name_list[1:]):
        cell = booksheet.cell(row=i+2, column=cnt)
        if name in name_sets['C']:
            cell.value = "已回家"
            cell.fill = sty.PatternFill(fill_type='solid',fgColor="ffff00")
        elif name in name_sets['B']:
            cell.value = "微信确认"
        elif name not in undo:
            cell.value = "微信确认"
        else:
            cell.value = "未确认"
            cell.fill = sty.PatternFill(fill_type='solid',fgColor="ff0000")
    workbook.save(path)

if __name__ == "__main__":

    w, q = get_map('数据表.xlsx')
    n = check_talbe('2256106_seg_1.csv', w, q)
    clock_in('2018暑期住宿.xlsx', n)
